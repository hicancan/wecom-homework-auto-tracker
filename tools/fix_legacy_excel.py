"""Temporary: fix empty 提交内容 cells in legacy collection Excel files.

Call this before running the interactive collector.
Delete this script and its call in scripts/run_extract_interactive.cmd
once all legacy collection tables have been replaced.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

FIXES = {
    "算法分析与设计作业[B240401-03][大二下]": "作业(.doc/.docx)",
    "算法分析与设计实验[B240401-03][大二下]": "实验报告(.doc/.docx)",
    "人工智能导论及其Python应用实践实验[B240402][大二下]": "实验报告(.doc/.docx)",
}


def main():
    config_dir = Path(__file__).resolve().parents[1] / "config"
    fixed = 0

    for filename, default in FIXES.items():
        excel_path = config_dir / f"{filename}.xlsx"
        if not excel_path.exists():
            continue

        wb = load_workbook(excel_path)
        ws = wb.active

        # Find the 提交内容 column
        content_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and "提交内容" in str(cell.value):
                content_col = col_idx
                break

        if content_col is None:
            print(f"  [SKIP] {filename}: no 提交内容 column")
            continue

        filled = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=content_col)
            if cell.value is None or str(cell.value).strip() in ("", "nan"):
                cell.value = default
                filled += 1

        if filled:
            wb.save(excel_path)
            print(f"  [OK] {filename}: filled {filled} cells with {default}")
            fixed += 1
        else:
            print(f"  [OK] {filename}: no empty cells")

    print(f"\nDone: fixed {fixed} file(s)")


if __name__ == "__main__":
    main()
